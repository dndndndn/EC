import hashlib
import json
import os

from django.core.paginator import Paginator
from django.http import StreamingHttpResponse, JsonResponse
from django.shortcuts import redirect
from django.shortcuts import render, HttpResponse
from django.urls import reverse
from openpyxl import load_workbook

from android import models as android_models
from . import forms
from . import models as login_models


# Create your views here.
# code:utf-8
def ins(request):
    return redirect(reverse('login'))


def index(request):
    user = login_models.User.objects.get(name=request.session['username'])
    if user.auth == 'admin' or user.auth == 'superadmin':
        admin = True
    return render(request, 'login/index.html', locals())


def hash_code(s, salt='mysite'):
    h = hashlib.sha256()
    s += salt
    h.update(s.encode())
    return h.hexdigest()


def login(request):
    if request.session.get('is_login', None):
        return redirect("/admin/")
    login_form = forms.UserForm()
    if request.method == "POST":
        login_form = forms.UserForm(request.POST)
        message = "请检查填写的内容！"
        if login_form.is_valid():
            username = login_form.cleaned_data['username']
            password = login_form.cleaned_data['password']
            try:
                user = login_models.User.objects.get(name=username)
                if user.password == hash_code(password):
                    request.session['is_login'] = True
                    request.session['user_id'] = user.id
                    request.session['username'] = user.name
                    return redirect(reverse('index'))
                else:
                    message = "密码错误"
            except:
                message = "用户不存在"
    return render(request, 'login/login.html', locals())


def register(request):
    if request.session.get('is_login', None):
        return redirect(reverse('admin'))
    register_form = forms.RegisterForm()
    if request.method == "POST":
        register_form = forms.RegisterForm(request.POST)
        message = "请检查填写内容"
        if register_form.is_valid():
            username = register_form.cleaned_data['username']
            password1 = register_form.cleaned_data['password1']
            password2 = register_form.cleaned_data['password2']
            email = register_form.cleaned_data['email']
            sex = register_form.cleaned_data['sex']
            student_id = register_form.cleaned_data['student_id']
            if password1 != password2:
                message = "两次输入密码不同"
                return render(request, 'login/register.html', locals())
            else:
                same_name_user = login_models.User.objects.filter(name=username)
                if same_name_user:
                    massage = "用户已经存在，请重新选择用户名"
                    return render(request, 'login/register.html', locals())
                same_email_user = login_models.User.objects.filter(email=email)
                if same_email_user:
                    message = "该邮箱地址已被注册，请使用别的邮箱"
                    return render(request, 'login/register.html', locals())

                # everything ok,then creat new account

                new_user = login_models.User()
                new_user.name = username
                new_user.password = hash_code(password1)
                new_user.student_id = student_id
                new_user.email = email
                new_user.sex = sex
                new_user.save()
                return redirect(reverse('login'))
    return render(request, 'login/register.html', locals())


def logout(request):
    if not request.session.get('is_login', None):
        return redirect(reverse('login'))
    request.session.flush()
    return redirect(reverse('login'))


from django.core.cache import cache
from django.http import HttpRequest
from django.utils.cache import get_cache_key


def expire_page(path):
    request = HttpRequest()
    request.path = path
    key = get_cache_key(request)
    if cache.has_key(key):
        cache.delete(key)


def admin(request):
    return redirect(reverse('account_all', args=(0, 1)))


def admin_account(request):
    return redirect(reverse('account_all', args=(0, 1)))


def admin_account_all(request, tag, page):
    if request.method == 'GET':
        user_list = login_models.User.objects.all()
        paginator = Paginator(user_list, 25)  # Show 25 contacts per page
        user = paginator.get_page(page)
        return render(request, 'login/admin_account_all.html', locals())
    elif request.method == 'POST':
        a = bytes.decode(request.body)
        b = a.split(',')
        for r in b:
            user = login_models.User.objects.get(student_id=r)
            user.delete()
        return JsonResponse(json.dumps({"success": "y"}), safe=False)


def admin_account_update(request, num):
    if request.method == 'GET':
        user = login_models.User.objects.get(student_id=num)

        return render(request, 'login/admin_account_update.html', locals())
    elif request.method == 'POST':
        user = login_models.User.objects.get(student_id=num)
        data = json.loads(request.body)
        print(data)
        user.student_id = data['student_id']
        user.name = data['name']
        user.school = data['school']
        user.email = data['email']
        user.password = hash_code(data['password'])
        user.sex = data['sex']
        print(user.student_id,
              user.name,
              user.school,
              user.email,
              user.password,
              user.school)
        user.save()
        return JsonResponse(json.dumps({"success": "y"}), safe=False)


def admin_account_groups(request):
    user = login_models.User.objects.all()
    return render(request, 'login/admin_account_all.html', locals())


def admin_account_upload(request):
    if request.method == 'GET':
        user = login_models.User.objects.all()
        return render(request, 'login/admin_account_upload.html', locals())
    elif request.method == 'POST':
        print(request.POST.get('target'))
        if request.POST.get('target') == 'new':
            user = login_models.User()
            data = request.POST
            print(data)
            user.student_id = data['student_id']
            user.name = data['name']
            user.school = data['school']
            user.email = data['email']
            user.password = hash_code(data['password'])
            user.sex = data['sex']
            print(user.student_id,
                  user.name,
                  user.school,
                  user.email,
                  user.password,
                  user.school)
            user.save()
            return JsonResponse(json.dumps({"success": "y"}), safe=False)
        if request.POST.get('target') == 'upload':
            obj = request.FILES.get('fafafa')
            file_path = os.path.join('store/upload', obj.name)
            f = open(file_path, 'wb')
            for chunk in obj.chunks():
                f.write(chunk)
            f.close()
            response = HttpResponse(json.dumps('shangch'))
            wb = load_workbook(r"G:\server\EC\store\upload\用户.xlsx")
            sheet = wb.get_sheet_by_name("Sheet1")
            for row in range(2, sheet.max_row + 1):
                q = login_models.User(
                    student_id=sheet['A' + str(row)].value,
                    name=sheet['B' + str(row)].value,
                    password=hash_code(sheet['G' + str(row)].value),
                    email=sheet['C' + str(row)].value,
                    sex=sheet['D' + str(row)].value,
                    auth=sheet['F' + str(row)].value,
                    school=sheet['E' + str(row)].value
                )
                q.save()
    return JsonResponse(json.dumps({"success": "n"}), safe=False)


# TODO：概要文档
# TODO：搜集信息
def account_download(request):
    file_path = r'G:\server\EC\store\download\用户表.xlsx'
    file_name = 'as.xlsx'

    def file_itertor(file_name, chunk_size=512):
        with open(file_name, 'rb') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break

    response = StreamingHttpResponse(file_itertor(file_path))
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(file_name)
    return response


def admin_question(request):
    return redirect((reverse('question_all', args=('0000', '0001'))))


def admin_question_all(request, tag, page):
    if request.method == 'GET':
        question_list = android_models.Question.objects.all()
        paginator = Paginator(question_list, 25)  # Show 25 contacts per page
        question = paginator.get_page(page)
        return render(request, 'login/admin_question_all.html', locals())
    elif request.method == 'POST':
        a = bytes.decode(request.body)
        b = a.split(',')
        for r in b:
            question = android_models.Question.objects.get(ID=r)
            question.delete()
        return JsonResponse(json.dumps({"success": "y"}), safe=False)


def admin_question_update(request, num):
    if request.method == 'GET':
        question = android_models.Question.objects.get(ID=num)
        choice_length = question.choice_length
        choice_list = question.choice.all()
        print(choice_list.first().ID)
        return render(request, 'login/admin_question_update.html', locals())
    elif request.method == 'POST':
        question = android_models.Question.objects.get(ID=num)
    return render(request, 'login/admin_question_update.html', locals())


def admin_question_groups(request, num):
    if request.method == 'GET':
        groups_list = android_models.QuestionGroup.objects.all()
        paginator = Paginator(groups_list, 25)  # Show 25 contacts per page
        groups = paginator.get_page(num)
        return render(request, 'login/adimin_question_groups.html', locals())
    elif request.method == 'POST':
        a = bytes.decode(request.body)
        b = a.split(',')
        for r in b:
            question = android_models.QuestionGroup.objects.get(ID=r)
            question.delete()
        return JsonResponse(json.dumps({"success": "y"}), safe=False)


def admin_question_groups_new(request):
    # todo:new groups
    if request.method == "GET":
        outquestion = android_models.Question.objects.all()
    elif request.method == "POST":
        print(request.POST)
        group = android_models.QuestionGroup()
        group.name = request.POST.get('groups_name')
        group.tips = request.POST.get('Tips')
        group.save()
        return JsonResponse(json.dumps({"success": "y"}), safe=False)
    return render(request, 'login/admin_question_groups_new.html', locals())


def admin_question_groups_update(request, ID):
    # todo:new groups
    if request.method == "GET":
        questionGroups = android_models.QuestionGroup.objects.get(ID=ID)
        print(questionGroups)
        inquestion = questionGroups.members.all()
        print(inquestion)
        outquestion = android_models.Question.objects.exclude(questiongroup__ID=ID)
        for i in outquestion:
            print(i.ID)
        print(outquestion)
    elif request.method == "POST":
        questionGroups = android_models.QuestionGroup.objects.get(ID=ID)

        IDarray = request.POST.get('groupsID')
        print(request.POST, IDarray)

        for i in IDarray:
            if i != ',':
                try:
                    questionGroups.members.get(ID=i)
                    continue
                except:
                    question = android_models.Question.objects.get(ID=i)
                    membership = android_models.QuestionMembership(question=question, group=questionGroups)
                    membership.save()
                    print(membership)
                    questionGroups.count += 1
        questionGroups.save()
        inquestion = questionGroups.members.all()
        for i in IDarray:
            if i != ',':
                try:
                    questionGroups.members.get(ID=i)
                    inquestion = inquestion.exclude(ID=i)
                    print(inquestion, 5)
                    continue
                except:
                    pass
        print(inquestion, 6)
        for i in inquestion:
            membership = android_models.QuestionMembership.objects.get(question=i, group=questionGroups)
            questionGroups.count -= 1
            membership.delete()
        questionGroups.save()
        return JsonResponse(json.dumps({"success": "y"}), safe=False)
    return render(request, 'login/admin_question_groups_update.html', locals())


def admin_question_upload(request):
    user = login_models.User.objects.all()
    if request.method == 'GET':
        user = login_models.User.objects.all()
        return render(request, 'login/admin_question_upload.html', locals())
    elif request.method == 'POST':
        print(request.POST, request.FILES)
        if request.POST.get('target') == "fileupload":
            response = HttpResponse(json.dumps('文件提交成功'))
        if request.POST.get('target') == "formupload":
            response = HttpResponse(json.dumps('表单提交成功'))
            question = android_models.Question()
            question.Q_Status = 0
            question.text = ""
            question.choice_length = request.POST.get("choice_length")
            question.general_priority = 2
            choicetype = 0
            for i in range(0, int(request.POST.get('choice_length'))):
                if request.POST.get("choice" + str(i)) == "true":
                    choicetype += 1
            if choicetype > 1:
                question.choice_type = 1
            else:
                question.choice_type = 0
            question.save()

            def list_input(length, name, new, question):
                list = []
                print(length)
                if length == 0:
                    return None
                for i in range(0, length):

                    list.append(new())
                    if request.POST.get("img" + name + str(i)):
                        list[i].type = "img"
                        obj = request.FILES.get('img' + name + str(i))
                        list[i].img = obj
                    if request.POST.get("txt" + name + str(i)):
                        list[i].type = "txt"
                        obj = request.POST.get("txt" + name + str(i))
                        list[i].text = obj
                    list[i].question = question
                    list[i].save()
                    if i != 0:
                        list[i - 1].next_content = list[i]
                        list[i - 1].save()
                list[len(list) - 1].next_content = None
                list[len(list) - 1].save()
                return list[0]

            question.stem = list_input(int(request.POST.get('stem_length')), "stem", android_models.Stem, question)
            question.solution = list_input(int(request.POST.get('solu_length')), "solu", android_models.Solution,
                                           question)
            question.save()
            question.choice_length = int(request.POST.get('choice_length'))
            for i in range(0, question.choice_length):
                choice = android_models.Choice()
                choice.related_question = question
                if request.POST.get('choice' + str(i)) == 'true':
                    choice.Bool = True
                else:
                    choice.Bool = False
                if request.POST.get('txt' + 'choice' + str(i)):
                    choice.type = 'txt'
                    choice.text = request.POST.get('txt' + 'choice' + str(i))
                elif request.POST.get('img' + 'choice' + str(i)):
                    choice.type = 'img'
                    choice.img = request.FILES.get('img' + 'choice' + str(i))
                choice.save()

                choicetips = android_models.ChoiceTips()
                choicetips.related_choice = choice
                if request.POST.get('txt' + 'choice_tips' + str(i)):
                    choicetips.type = 'txt'
                    choicetips.text = request.POST.get('txt' + 'choice_tips' + str(i))
                elif request.POST.get('img' + 'choice_tips' + str(i)):
                    choicetips.type = 'img'
                    choicetips.img = request.FILES.get('img' + 'choice_tips' + str(i))
                else:
                    choicetips.type = 'null'
                choicetips.save()
    return response


def admin_question_search(request):
    user = login_models.User.objects.all()
    # redirect to all
    return render(request, 'login/admin_question_all.html', locals())


def question_download(request):
    file_path = r'G:\server\EC\store\download\用户表.xlsx'
    file_name = 'as.xlsx'

    def file_itertor(file_name, chunk_size=512):
        with open(file_name, 'rb') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break

    response = StreamingHttpResponse(file_itertor(file_path))
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(file_name)
    print(response)
    return response


def admin_test(request):
    user = login_models.User.objects.all()
    return render(request, 'login/admin_test.html', locals())


def admin_forum(request):
    user = login_models.User.objects.all()
    return render(request, 'login/admin_forum.html', locals())


def admin_syllabus(request):
    user = login_models.User.objects.all()
    return render(request, 'login/admin_syllabus.html', locals())


def admin_homework(request):
    user = login_models.User.objects.all()
    return render(request, 'login/admin_homework.html', locals())


def admin_feedback(request):
    user = login_models.User.objects.all()
    return render(request, 'login/admin_feedback.html', locals())


def user_filter(request):
    return redirect(reverse('register'))


def user_detail(request, id, name):
    a = login_models.User.objects.get(student_id=id)
    return render(request, "login/user_detail.html", locals())


def page_not_found(request):
    return render(request, 'login/404.html')


def page_error(request):
    return render(request, 'login/500.html')


def permission_denied(request):
    return render(request, 'login/403.html')


def bad_request(request):
    return render(request, 'login/400.html')
